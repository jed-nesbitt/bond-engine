from __future__ import annotations

import argparse
import getpass
import json
import platform
import socket
import time
from datetime import datetime, timezone
from pathlib import Path
from uuid import uuid4

import pandas as pd

from src.batch import run_batch


def _safe_weighted_avg(values: pd.Series, weights: pd.Series) -> float:
    w = weights.fillna(0.0)
    v = values.fillna(0.0)
    denom = float(w.sum())
    return float((v * w).sum() / denom) if denom > 0 else float("nan")


def _make_run_dir(out_root: Path) -> tuple[Path, str]:
    """
    Creates: <out_root>/runs/<YYYYMMDD_HHMMSS>_<8charid>/
    Returns (run_dir, run_id)
    """
    ts = datetime.now().astimezone().strftime("%Y%m%d_%H%M%S")
    run_id = f"{ts}_{uuid4().hex[:8]}"
    run_dir = out_root / "runs" / run_id
    run_dir.mkdir(parents=True, exist_ok=False)
    return run_dir, run_id


def _write_run_metadata(run_dir: Path, metadata: dict) -> Path:
    path = run_dir / "run_metadata.json"
    with path.open("w", encoding="utf-8") as f:
        json.dump(metadata, f, indent=2, sort_keys=True, default=str)
    return path


def _resolve_input_path(raw: str) -> Path:
    """
    If --input is a filename (or relative path) and does not exist as given,
    try resolving it under ./input/
    """
    p = Path(raw)

    # If path exists as provided, use it
    if p.exists():
        return p

    # If user passed something like "bonds.csv", try "input/bonds.csv"
    candidate = Path("input") / p
    if candidate.exists():
        return candidate

    # Otherwise return original (so we can raise a clean error)
    return p


def main() -> None:
    parser = argparse.ArgumentParser(description="Bond batch analytics: CSV in, stats CSV/XLSX out.")
    parser.add_argument(
        "--input",
        type=str,
        default=str(Path("input") / "example_bonds.csv"),
        help="Path to input CSV (defaults to input/example_bonds.csv). If not found, will try input/<value>.",
    )
    parser.add_argument("--outdir", type=str, default="output", help="Base output folder (runs go into outdir/runs/...)")
    args = parser.parse_args()

    out_root = Path(args.outdir)
    out_root.mkdir(parents=True, exist_ok=True)

    run_dir, run_id = _make_run_dir(out_root)

    t0 = time.time()
    start_utc = datetime.now(timezone.utc).isoformat()
    start_local = datetime.now().astimezone().isoformat()

    input_csv = _resolve_input_path(args.input)

    metadata: dict = {
        "run_id": run_id,
        "status": "running",
        "start_utc": start_utc,
        "start_local": start_local,
        "input_arg": args.input,
        "input_resolved": str(input_csv),
        "out_root": str(out_root),
        "run_dir": str(run_dir),
        "environment": {
            "python_version": platform.python_version(),
            "platform": platform.platform(),
            "hostname": socket.gethostname(),
            "user": getpass.getuser(),
        },
        "outputs": {},
        "counts": {},
    }

    try:
        if not input_csv.exists():
            raise FileNotFoundError(
                f"Input CSV not found: {input_csv}\n"
                f"Tried: '{args.input}' and, if needed, 'input/{Path(args.input).name}'"
            )

        res = run_batch(input_csv)

        out_csv = run_dir / "bond_statistics.csv"
        out_xlsx = run_dir / "bond_statistics.xlsx"

        res.to_csv(out_csv, index=False)

        # ---- Build Summary (only from OK rows) ----
        ok = res[res["status"] == "OK"].copy()

        if len(ok) == 0:
            summary_main = pd.DataFrame(
                {"metric": ["OK_rows", "ERROR_rows"], "value": [0, int((res["status"] == "ERROR").sum())]}
            )
            by_day = pd.DataFrame()
            by_freq = pd.DataFrame()
            top_risk = pd.DataFrame()
        else:
            # Ensure portfolio fields exist (for backward compatibility)
            if "quantity" not in ok.columns:
                ok["quantity"] = 1.0
            if "mv_dirty" not in ok.columns:
                ok["mv_dirty"] = ok["dirty_price"] * ok["quantity"]
            if "mv_clean" not in ok.columns:
                ok["mv_clean"] = ok["clean_price"] * ok["quantity"]
            if "accrued_interest_total" not in ok.columns:
                ok["accrued_interest_total"] = ok["accrued_interest"] * ok["quantity"]
            if "dv01_total" not in ok.columns:
                ok["dv01_total"] = ok["dv01_per_100"] * ok["quantity"]

            total_mv_dirty = float(ok["mv_dirty"].sum())
            total_mv_clean = float(ok["mv_clean"].sum())
            total_ai = float(ok["accrued_interest_total"].sum())
            total_dv01 = float(ok["dv01_total"].sum())

            w = ok["mv_dirty"]  # weight by dirty market value

            summary_main = pd.DataFrame(
                {
                    "metric": [
                        "OK_rows",
                        "ERROR_rows",
                        "Total_MV_Dirty",
                        "Total_MV_Clean",
                        "Total_Accrued_Interest",
                        "Weighted_YTM",
                        "Weighted_Modified_Duration_Years",
                        "Weighted_Macaulay_Duration_Years",
                        "Total_DV01",
                        "Weighted_Convexity_Num",
                        "Weighted_TTM_Years",
                    ],
                    "value": [
                        int(len(ok)),
                        int((res["status"] == "ERROR").sum()),
                        total_mv_dirty,
                        total_mv_clean,
                        total_ai,
                        _safe_weighted_avg(ok["ytm"], w),
                        _safe_weighted_avg(ok["modified_duration_years"], w),
                        _safe_weighted_avg(ok["macaulay_duration_years"], w),
                        total_dv01,
                        _safe_weighted_avg(ok["convexity_num"], w),
                        _safe_weighted_avg(ok["time_to_maturity_years"], w),
                    ],
                }
            )

            # Breakdown tables
            by_day = (
                ok.groupby("day_count", dropna=False)
                .agg(
                    bonds=("bond_id", "count"),
                    mv_dirty=("mv_dirty", "sum"),
                    weighted_ytm=("ytm", lambda s: _safe_weighted_avg(s, ok.loc[s.index, "mv_dirty"])),
                    weighted_mod_dur=("modified_duration_years", lambda s: _safe_weighted_avg(s, ok.loc[s.index, "mv_dirty"])),
                    total_dv01=("dv01_total", "sum"),
                )
                .reset_index()
                .sort_values("mv_dirty", ascending=False)
            )

            by_freq = (
                ok.groupby("frequency", dropna=False)
                .agg(
                    bonds=("bond_id", "count"),
                    mv_dirty=("mv_dirty", "sum"),
                    weighted_ytm=("ytm", lambda s: _safe_weighted_avg(s, ok.loc[s.index, "mv_dirty"])),
                    weighted_mod_dur=("modified_duration_years", lambda s: _safe_weighted_avg(s, ok.loc[s.index, "mv_dirty"])),
                    total_dv01=("dv01_total", "sum"),
                )
                .reset_index()
                .sort_values("mv_dirty", ascending=False)
            )

            # Top risk contributors (by DV01 absolute)
            top_risk = (
                ok.assign(dv01_abs=ok["dv01_total"].abs())
                .sort_values("dv01_abs", ascending=False)
                .loc[
                    :,
                    [
                        "bond_id",
                        "quantity",
                        "mv_dirty",
                        "ytm",
                        "modified_duration_years",
                        "dv01_total",
                        "dirty_price",
                        "clean_price",
                    ],
                ]
                .head(10)
            )

        # ---- Write Excel with two tabs ----
        with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
            res.to_excel(writer, sheet_name="Bonds", index=False)

            # Summary tab with multiple blocks
            summary_main.to_excel(writer, sheet_name="Summary", index=False, startrow=0)

            r = len(summary_main) + 3
            if len(by_day) > 0:
                pd.DataFrame({"Breakdown_by_DayCount": []}).to_excel(writer, sheet_name="Summary", index=False, startrow=r - 1)
                by_day.to_excel(writer, sheet_name="Summary", index=False, startrow=r)
                r += len(by_day) + 3

            if len(by_freq) > 0:
                pd.DataFrame({"Breakdown_by_Frequency": []}).to_excel(writer, sheet_name="Summary", index=False, startrow=r - 1)
                by_freq.to_excel(writer, sheet_name="Summary", index=False, startrow=r)
                r += len(by_freq) + 3

            if len(top_risk) > 0:
                pd.DataFrame({"Top_10_by_Absolute_DV01": []}).to_excel(writer, sheet_name="Summary", index=False, startrow=r - 1)
                top_risk.to_excel(writer, sheet_name="Summary", index=False, startrow=r)

            # -----------------------------
            # Formatting (openpyxl)
            # -----------------------------
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter

            header_font = Font(bold=True)
            title_font = Font(bold=True)
            center = Alignment(horizontal="center")

            def is_blank(v: object) -> bool:
                return v is None or (isinstance(v, str) and v.strip() == "")

            def auto_width(ws, min_w: int = 10, max_w: int = 45) -> None:
                for col in range(1, ws.max_column + 1):
                    max_len = 0
                    for row in range(1, ws.max_row + 1):
                        val = ws.cell(row=row, column=col).value
                        if val is None:
                            continue
                        s = str(val)
                        max_len = max(max_len, len(s))
                    ws.column_dimensions[get_column_letter(col)].width = max(min_w, min(max_w, max_len + 2))

            def set_table_header_style(ws, header_row: int, end_col: int) -> None:
                for c in range(1, end_col + 1):
                    cell = ws.cell(row=header_row, column=c)
                    cell.font = header_font
                    cell.alignment = center

            def apply_formats_by_headers(ws, header_row: int, end_row: int) -> None:
                date_cols = {"issue_date", "settle_date", "maturity_date"}
                pct_cols = {"coupon_rate", "ytm", "input_ytm", "weighted_ytm"}
                price_cols_4dp = {"clean_price", "dirty_price", "accrued_interest", "price_change_+25bp", "price_change_+100bp"}
                money_cols_2dp = {"mv_dirty", "mv_clean", "accrued_interest_total", "Total_MV_Dirty", "Total_MV_Clean", "Total_Accrued_Interest"}
                dv01_cols = {"dv01_per_100", "dv01_total", "total_dv01", "Total_DV01"}
                dur_cols = {
                    "macaulay_duration_years",
                    "modified_duration_years",
                    "weighted_mod_dur",
                    "weighted_mod_dur_years",
                    "Weighted_Modified_Duration_Years",
                    "Weighted_Macaulay_Duration_Years",
                    "time_to_maturity_years",
                    "Weighted_TTM_Years",
                }
                conv_cols = {"convexity_num", "Weighted_Convexity_Num"}
                qty_cols = {"quantity", "bonds"}

                headers: dict[str, int] = {}
                col = 1
                while col <= ws.max_column:
                    h = ws.cell(row=header_row, column=col).value
                    if is_blank(h):
                        break
                    headers[str(h)] = col
                    col += 1

                for h, c in headers.items():
                    h_norm = h.strip()
                    if h_norm in date_cols:
                        fmt = "yyyy-mm-dd"
                    elif h_norm in pct_cols or "YTM" in h_norm:
                        fmt = "0.00%"
                    elif h_norm in qty_cols:
                        fmt = "0.####"
                    elif h_norm in dv01_cols:
                        fmt = "0.0000"
                    elif h_norm in dur_cols:
                        fmt = "0.0000"
                    elif h_norm in conv_cols:
                        fmt = "0.0000"
                    elif h_norm in money_cols_2dp:
                        fmt = "#,##0.00"
                    elif h_norm in price_cols_4dp:
                        fmt = "0.0000"
                    else:
                        continue

                    for rr in range(header_row + 1, end_row + 1):
                        ws.cell(row=rr, column=c).number_format = fmt

            def format_table_at(ws, header_row: int, start_col: int = 1) -> None:
                end_col = start_col
                while end_col <= ws.max_column and not is_blank(ws.cell(row=header_row, column=end_col).value):
                    end_col += 1
                end_col -= 1
                if end_col < start_col:
                    return

                end_row = header_row
                rr = header_row + 1
                while rr <= ws.max_row:
                    row_blank = True
                    for cc in range(start_col, end_col + 1):
                        if not is_blank(ws.cell(row=rr, column=cc).value):
                            row_blank = False
                            break
                    if row_blank:
                        break
                    end_row = rr
                    rr += 1

                set_table_header_style(ws, header_row, end_col)
                apply_formats_by_headers(ws, header_row, end_row)

            def find_row_in_col_a(ws, text: str) -> int | None:
                for rr in range(1, ws.max_row + 1):
                    v = ws.cell(row=rr, column=1).value
                    if isinstance(v, str) and v.strip() == text:
                        return rr
                return None

            def set_autofilter(ws, header_row: int = 1) -> None:
                last_row = ws.max_row
                last_col = ws.max_column
                if last_row >= header_row and last_col >= 1:
                    ref = f"A{header_row}:{get_column_letter(last_col)}{last_row}"
                    ws.auto_filter.ref = ref

            ws_bonds = writer.book["Bonds"]
            ws_bonds.freeze_panes = "A2"
            set_autofilter(ws_bonds, header_row=1)
            format_table_at(ws_bonds, header_row=1)
            auto_width(ws_bonds)

            ws_sum = writer.book["Summary"]
            ws_sum.freeze_panes = "A2"
            format_table_at(ws_sum, header_row=1)

            for title in ("Breakdown_by_DayCount", "Breakdown_by_Frequency", "Top_10_by_Absolute_DV01"):
                title_row = find_row_in_col_a(ws_sum, title)
                if title_row is None:
                    continue
                ws_sum.cell(row=title_row, column=1).font = title_font
                format_table_at(ws_sum, header_row=title_row + 1)

            auto_width(ws_sum)

        metadata["status"] = "success"
        metadata["outputs"] = {
            "bond_statistics_csv": str(out_csv),
            "bond_statistics_xlsx": str(out_xlsx),
        }
        metadata["counts"] = {
            "rows_total": int(len(res)),
            "rows_ok": int((res["status"] == "OK").sum()),
            "rows_error": int((res["status"] == "ERROR").sum()),
        }

        print(f"\nRun folder: {run_dir}")
        print(f"✅ Wrote: {out_csv}")
        print(f"✅ Wrote: {out_xlsx}")
        print(f"Rows: {len(res)} | OK: {(res['status']=='OK').sum()} | ERROR: {(res['status']=='ERROR').sum()}")

    except Exception as e:
        metadata["status"] = "error"
        metadata["error"] = {"type": type(e).__name__, "message": str(e)}
        raise

    finally:
        metadata["end_utc"] = datetime.now(timezone.utc).isoformat()
        metadata["duration_seconds"] = round(time.time() - t0, 3)
        meta_path = _write_run_metadata(run_dir, metadata)
        print(f"📝 Saved run metadata: {meta_path}")


if __name__ == "__main__":
    main()
